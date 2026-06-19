import os
import openpyxl
import random

def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(base_dir, 'dataset_smart_souvenir.xlsx')
    
    print("Loading database Excel from:", excel_path)
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} tidak ditemukan!")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    
    # Store existing rows
    existing_headers = rows[:3]
    existing_data = []
    seen_products = set()
    
    for r in rows[3:]:
        nama_produk = r[0]
        perusahaan = r[1]
        kategori = r[2]
        if nama_produk is not None and kategori is not None:
            existing_data.append((nama_produk, perusahaan, kategori))
            seen_products.add((str(nama_produk).strip().lower(), str(kategori).strip().lower()))
            
    print(f"Total existing valid data: {len(existing_data)}")
    
    # Category configurations for generation
    DATA_TEMPLATES = {
        'Alat Tulis & Kosmetik': {
            'products': ['Pensil 2B', 'Pulpen Gel', 'Buku Tulis 38 Lembar', 'Buku Tulis 58 Lembar', 'Penghapus Karet', 'Penggaris Besi 30cm', 'Rautan Pensil', 'Sunscreen SPF 50', 'Matte Lip Cream', 'Micellar Water', 'Bedak Tabur', 'Mascara Waterproof', 'Pensil Alis', 'Eyeliner Black', 'Highlighter', 'Blush On', 'Foundation Liquid', 'Marker Permanent', 'Crayon 12 Warna', 'Buku Gambar A4', 'Lem Kertas Cair', 'Gunting Kecil', 'Double Tape', 'Lip Balm Balm', 'Facial Mist Hydrating', 'Clay Mask Green Tea', 'Serum Brightening Vitamin C', 'Pulpen Cetek', 'Pensil Warna 24', 'Buku Agenda', 'Cushion Glow', 'Lip Tint Red', 'Eyebrow Gel', 'Setting Spray'],
            'brands': ['Joyko', 'Kenko', 'Faber-Castell', 'Wardah', 'Emina', 'Pixy', 'Ponds', 'Garnier', 'Viva', 'Purbasari', 'Make Over', 'Kahf', 'Focallure', 'Snowman', 'Paperline', 'Pilot', 'Pentel', 'Mirabella', 'Sariayu', 'Kino', 'Somethinc', 'Skintific'],
            'companies': ['PT Joyko Makmur', 'PT Kenko Indonesia', 'PT Faber-Castell International Indonesia', 'PT Paragon Technology and Innovation', 'PT Unilever Indonesia Tbk', 'PT L\'Oreal Indonesia', 'PT Lion Wings', 'PT Martina Berto Tbk', 'PT Kino Indonesia Tbk', 'PT Somethinc Indonesia', 'PT Skintific Beauty']
        },
        'Bahan Makanan': {
            'products': ['Minyak Goreng Sawit', 'Tepung Terigu', 'Mentega Serbaguna', 'Kecap Manis', 'Saus Sambal Ekstra Pedas', 'Penyedap Rasa Ayam', 'Penyedap Rasa Sapi', 'Garam Dapur Beryodium', 'Gula Pasir Putih', 'Santan Kelapa Cair', 'Merica Bubuk', 'Bumbu Nasi Goreng Instan', 'Kaldu Jamur Bubuk', 'Margarin Lezat', 'Cuka Dapur', 'Minyak Wijen', 'Kecap Asin', 'Saus Tiram', 'Bumbu Rendang Instan', 'Tepung Tapioka', 'Tepung Beras', 'Gula Merah Jawa', 'Minyak Goreng Kelapa', 'Saus Tomat Botol', 'Tepung Maizena', 'Ragi Instan', 'Bumbu Opor Instan', 'Kecap Inggris'],
            'brands': ['Indofood', 'Sasa', 'Ajinomoto', 'Koepoe Koepoe', 'Blue Band', 'Filma', 'Bimoli', 'Sunco', 'Rose Brand', 'Segitiga Biru', 'Kunci Biru', 'Cakra Kembar', 'Bango', 'ABC', 'Sedaap', 'Sania', 'Tropica', 'Minyak Kita', 'Knorr', 'Ladaku'],
            'companies': ['PT Indofood CBP Sukses Makmur Tbk', 'PT Sasa Inti', 'PT Ajinomoto Indonesia', 'PT Gunawan Fajar', 'PT Upfield Foods Indonesia', 'PT Salim Ivomas Pratama Tbk', 'PT Wilmar Cahaya Indonesia Tbk', 'PT Heinz ABC Indonesia', 'PT Wings Surya', 'PT Unilever Indonesia Tbk']
        },
        'Kebutuhan Bayi': {
            'products': ['Baby Powder Active Fresh', 'Baby Bath Hair & Body', 'Baby Oil Aloe Vera', 'Minyak Telon Plus', 'Tissue Basah Baby Wipes', 'Popok Celana', 'Diapers Baby Pants', 'Botol Susu Baby', 'Sabun Cuci Botol Bayi', 'Cotton Bud Baby', 'Baby Lotion Soft', 'Baby Cologne Fresh', 'Bubur Bayi Instan', 'Susu Formula Infant', 'Minyak Kayu Putih Bayi', 'Bedak Padat Bayi', 'Baby Hair Lotion', 'Baby Sunscreen Cream', 'Teether Silicone', 'Bubur Tim Beras Merah', 'Susu Pertumbuhan Balita'],
            'brands': ['Zwitsal', 'Mitu', 'My Baby', 'Cussons Baby', 'Pigeon', 'Johnson\'s', 'Sweety', 'MamyPoko', 'Baby Happy', 'Merries', 'Sun', 'Milna', 'SGM', 'Dancow', 'Bebelac', 'Morinaga', 'Lactogrow'],
            'companies': ['PT Unilever Indonesia Tbk', 'PT PZ Cussons Indonesia', 'PT Pigeon Indonesia', 'PT Johnson & Johnson Indonesia', 'PT Sweety Indopaper', 'PT Unicharm Indonesia Tbk', 'PT Nestlé Indonesia', 'PT Sarihusada Generasi Mahardhika', 'PT Kalbe Farma Tbk']
        },
        'Makanan Ringan': {
            'products': ['Wafer Cokelat', 'Keripik Kentang Keju', 'Kacang Kulit Sangrai', 'Biskuit Kelapa', 'Snack Ekstrudat Keju', 'Chiki Balls Cokelat', 'Keripik Singkong Pedas', 'Cookies Choco Chip', 'Wafer Roll Vanilla', 'Nastar Keju', 'Snack Rumput Laut', 'Pilus Sapi Panggang', 'Keripik Tempe Orginal', 'Biskuit Marie', 'Popcorn Asin', 'Jelly Mangga', 'Permen Rasa Mint', 'Cokelat Bar', 'Snack Kentang Barbeque', 'Kacang Atom Pedas', 'Biskuit Sandwich Cokelat', 'Permen Karet'],
            'brands': ['Mayora', 'Nabisco', 'Garuda', 'Dua Kelinci', 'Khong Guan', 'Tango', 'Oishi', 'Gery', 'Nextar', 'Beng Beng', 'Chitato', 'Lays', 'Chiki', 'Cheetos', 'Kusuka', 'Silverqueen', 'Dilan', 'Oreo', 'Piatattos', 'Roma', 'Slai Olai'],
            'companies': ['PT Mayora Indah Tbk', 'PT Indofood CBP Sukses Makmur Tbk', 'PT GarudaFood Putra Putri Jaya Tbk', 'PT Dua Kelinci', 'PT Khong Guan Biscuit Factory Indonesia', 'PT Orang Tua Group', 'PT Liwayway Publishing', 'PT Kaldu Sari Nabati Indonesia', 'PT Dolphin Food & Beverage Industry', 'PT Mondelez Indonesia']
        },
        'Minuman': {
            'products': ['Air Mineral Botol', 'Teh Melati Botol', 'Susu UHT Full Cream', 'Kopi Susu Instan', 'Kopi Bubuk Murni', 'Minuman Isotonik', 'Susu Kental Manis', 'RTD Coffee Latte', 'Susu Steril Pouch', 'Jus Apel Kemasan Kotak', 'Jus Jeruk', 'Teh Hijau Celup', 'Minuman Soda Lemon', 'Air Demineral', 'Minuman Cokelat Malt', 'Yoghurt Drink Stroberi', 'Air Kelapa Kemasan', 'Teh Tarik Instan', 'RTD Milk Tea Botol', 'Jus Mangga Pouch', 'Kopi Hitam Sachet'],
            'brands': ['Aqua', 'Vit', 'Le Minerale', 'Teh Pucuk Harum', 'Ultra Jaya', 'Frisian Flag', 'Indomilk', 'Torabika', 'Kapal Api', 'Nescafe', 'Pocari Sweat', 'Sprite', 'Coca-Cola', 'Fanta', 'Teh Kotak', 'Milo', 'Cimory', 'Hydro Coco', 'Teh Botol Sosro', 'Kratingdaeng', 'Kopi Luwak'],
            'companies': ['PT Tirta Investama', 'PT Mayora Indah Tbk', 'PT Ultrajaya Milk Industry & Trading Company Tbk', 'PT Frisian Flag Indonesia', 'PT Indolakto', 'PT Santos Jaya Abadi', 'PT Nestlé Indonesia', 'PT Amerta Indah Otsuka', 'PT Coca-Cola Bottling Indonesia', 'PT Cimory Tbk', 'PT Sinar Sosro']
        },
        'Perawatan Diri': {
            'products': ['Sabun Mandi Cair', 'Pasta Gigi Herbal', 'Shampoo Anti Dandruff', 'Conditioner Hair Fall', 'Deodorant Roll On', 'Mouthwash Fresh Mint', 'Sabun Batang Antiseptik', 'Facial Foam Deep Clean', 'Pembersih Wajah Pria', 'Hand & Body Lotion', 'Pembersih Wajah Wanita', 'Hair Gel Strong', 'Sabun Wajah Jerawat', 'Pencukur Jenggot', 'Body Scrub Sakura', 'Sabun Mandi Cair Antibakteri', 'Pembersih Daerah Kewanitaan', 'Shampoo Hijab Fresh', 'Serum Rambut Hair Fall', 'Sabun Cuci Tangan Botol'],
            'brands': ['Pepsodent', 'Lifebuoy', 'Sunsilk', 'Clear', 'Dove', 'Rexona', 'Biore', 'Ciptadent', 'Kodomo', 'Colgate', 'Palmolive', 'Sensodyne', 'Gillette', 'Nivea', 'Vaseline', 'Dettol', 'Kahf', 'Mens Biore', 'Safi'],
            'companies': ['PT Unilever Indonesia Tbk', 'PT Kao Indonesia', 'PT Lion Wings', 'PT Colgate-Palmolive Indonesia', 'PT GlaxoSmithKline Indonesia', 'PT Procter & Gamble Home Products Indonesia', 'PT Beiersdorf Indonesia', 'PT Reckitt Benckiser Indonesia', 'PT Paragon Technology and Innovation']
        },
        'Perawatan Rumah': {
            'products': ['Detergen Bubuk', 'Sabun Cuci Piring Cair Lemon', 'Pewangi & Pelembut Pakaian', 'Karbol Wangi Pine', 'Pembersih Lantai Citrus', 'Pembersih Porselen Harpic', 'Cairan Penghilang Noda', 'Detergen Cair Konsentrat', 'Pembasmi Serangga Semprot', 'Pencegah Jamur Pakaian', 'Tisu Wajah Soft Pack', 'Tisu Toilet Roll', 'Pembersih Kaca Spray', 'Sabun Cuci Tangan Pouch', 'Pembersih Saluran Air', 'Pengharum Ruangan Gel', 'Tisu Basah Antiseptik Rumah'],
            'brands': ['So Klin', 'Daia', 'Ekonomi', 'Rinso', 'Sunlight', 'Molto', 'Wipol', 'Domestos', 'Harpic', 'Vanish', 'Baygon', 'Vape', 'Bagus', 'Paseo', 'Tessa', 'Mama Lemon', 'Glade', 'Stella'],
            'companies': ['PT Wings Surya', 'PT Unilever Indonesia Tbk', 'PT Reckitt Benckiser Indonesia', 'PT Johnson Home Products', 'PT Fumakilla Indonesia', 'PT Bagus Pandan Jaya', 'PT Pindo Deli Pulp & Paper Mills', 'PT SC Johnson Landmark', 'PT Godrej Consumer Products Indonesia']
        }
    }
    
    variants = [
        'Rasa Cokelat', 'Rasa Keju', 'Rasa Stroberi', 'Rasa Vanila', 'Rasa Melon', 'Rasa Mangga', 'Rasa Jeruk',
        'Ekstrak Lidah Buaya', 'Wangi Lavender', 'Wangi Rose', 'Wangi Lemon', 'Wangi Apple', 'Wangi Jasmine',
        'Kemasan Botol', 'Kemasan Sachet', 'Kemasan Pouch', 'Kemasan Kaleng', 'Kemasan Renceng', 'Kemasan Pack',
        'Ukuran 100g', 'Ukuran 200g', 'Ukuran 500g', 'Ukuran 1kg', 'Ukuran 250ml', 'Ukuran 500ml', 'Ukuran 1L',
        'Isi 10 Pcs', 'Isi 20 Pcs', 'Isi 50s', 'Formula Active', 'Ekstra Wangi', 'Anti Bacterial', 'Herbal Alami'
    ]
    
    # Target total rows
    target_rows = 5000
    current_count = len(existing_data)
    needed = target_rows - current_count
    
    print(f"Generating {needed} new unique rows...")
    categories_list = list(DATA_TEMPLATES.keys())
    
    generated_count = 0
    attempts = 0
    max_attempts = needed * 100
    
    new_rows = []
    
    while generated_count < needed and attempts < max_attempts:
        attempts += 1
        
        # Select random category
        cat = random.choice(categories_list)
        templates = DATA_TEMPLATES[cat]
        
        prod_base = random.choice(templates['products'])
        brand = random.choice(templates['brands'])
        comp = random.choice(templates['companies'])
        
        # Select 1 or 2 variants to add variety
        num_vars = random.choice([1, 2])
        vars_selected = random.sample(variants, num_vars)
        vars_str = " ".join(vars_selected)
        
        # Dynamic product name styling
        style = random.choice([1, 2, 3])
        if style == 1:
            prod_name = f"{brand} {prod_base} {vars_str}"
        elif style == 2:
            prod_name = f"{prod_base} {brand} {vars_str}"
        else:
            prod_name = f"{brand} {vars_str} {prod_base}"
            
        prod_name_clean = str(prod_name).strip()
        key = (prod_name_clean.lower(), cat.lower())
        
        if key not in seen_products:
            seen_products.add(key)
            new_rows.append((prod_name_clean, comp, cat))
            generated_count += 1
            
    print(f"Generated {len(new_rows)} new rows.")
    
    # Combine existing and new data
    final_data = existing_data + new_rows
    print(f"Total dataset size: {len(final_data)} rows.")
    
    # Write back to Excel
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = wb.active.title
    
    # Write headers
    for r_idx, h_row in enumerate(existing_headers, 1):
        for c_idx, val in enumerate(h_row, 1):
            new_ws.cell(r_idx, c_idx, val)
            
    # Write data
    for r_idx, data_row in enumerate(final_data, 4):
        for c_idx, val in enumerate(data_row, 1):
            new_ws.cell(r_idx, c_idx, val)
            
    new_wb.save(excel_path)
    print("dataset_smart_souvenir.xlsx successfully expanded to 5000 rows!")

if __name__ == '__main__':
    main()
